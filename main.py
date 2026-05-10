import argparse
import os
from datetime import datetime
from typing import Any

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment

from agent.orchestrator import ReportOrchestrator
import yaml


def _load_config(config_path: str) -> dict:
    with open(config_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def _ensure_sample_excel_exists(excel_path: str) -> None:
    if os.path.exists(excel_path):
        return

    os.makedirs(os.path.dirname(os.path.abspath(excel_path)), exist_ok=True)

    wb = Workbook()

    def set_percent(ws, cell, value_ratio):
        cell.value = float(value_ratio)
        cell.number_format = "0.00%"
        cell.alignment = Alignment(horizontal="center")

    # --- Cover Page --- 4 columns: Team | KPI Name | Definition | Sub-Definition
    ws_cover = wb.active
    ws_cover.title = "Cover Page"
    cover_rows = [
        # MIM
        ["MIM", "Total MI", "Total no. of Major Incidents reported in period", "Includes P1 and P2 incidents raised in ServiceNow"],
        ["", "MIR Conducted", "Major Incident Review conducted after each P1/P2", "Post-incident review to identify root cause and corrective actions"],
        ["", "# of P1", "Number of Priority 1 incidents in the period", "Critical incidents requiring immediate escalation to senior management"],
        ["", "# of P2", "Number of Priority 2 incidents in the period", "High priority incidents requiring prompt resolution within 8 hours"],
        ["", "100% Adherence to the MIM process (>99%)", "Every step of the MIM workflow is followed exactly as defined", "Includes logging, escalation, communication and closure steps"],
        ["", "Accurate stakeholder communications (>99%)", "Timely and accurate notifications sent to all relevant stakeholders", "Communication covers P1/P2 declaration, updates and resolution"],
        ["", "Consistency in the Service now fields (>99%)", "Correct categorization and sub-categorization in ServiceNow", "All mandatory incident fields populated correctly and consistently"],
        ["", "MIR Timeline (>99%)", "MIR completed within agreed timeframe after resolving MI", "MIR must be completed within 5 business days of incident closure"],
        ["", "SLA compliance P1/P2 (>95%)", "Percentage of P1/P2 incidents resolved within agreed SLA", "P1 target: resolve within 4 hrs; P2 target: resolve within 8 hrs"],
        ["", "SMS within 20 min P1/P2 (>99%)", "SMS alert sent within 20 minutes of P1/P2 declaration", "Automated SMS notification to on-call teams and stakeholders"],
        ["", "Email within 20 min engaged (>99%)", "Email notification sent within 20 minutes of team engagement", "Formal email to distribution list confirming team engaged"],
        ["", "Update cadence P1 30m P2 60m (>99%)", "Regular updates sent at defined cadence during incident", "P1: 30-minute updates; P2: 60-minute updates via Email & SMS"],
        ["", "MI whiteboard accuracy (>99%)", "All key incident details correctly recorded in real-time", "ServiceNow whiteboard fields kept current throughout the incident"],
        ["", "Effective handover for RCA (>99%)", "Complete handover of incident information to Problem Management", "Full RCA documentation and evidence passed to PM team at MIR"],
        # Change
        ["Change", "Total Changes", "Total number of change requests raised in the period", "Includes standard, normal and emergency change types"],
        ["", "Emergency Changes", "Number of emergency change requests raised", "Changes bypassing standard CAB approval due to urgency"],
        ["", "Change Success Ratio (>99%)", "Percentage of changes implemented successfully without incident", "Change passes CAB, executes within planned window, no rollback"],
        ["", "Change Causing MI (<5%)", "Percentage of changes that caused a Major Incident", "Changes that directly triggered a P1 or P2 incident post-implementation"],
        ["", "Failed Change Ratio (<1%)", "Percentage of changes that failed and required rollback", "Changes that could not complete and were rolled back to previous state"],
        ["", "Unauthorised Change rate (<1%)", "Percentage of changes implemented without proper CAB approval", "Emergency changes without timely post-implementation CAB review"],
        # Problem
        ["Problem", "Open Problems", "Total number of open problem records in ServiceNow", "All active problem records regardless of priority or age"],
        ["", "Problems Resolved", "Number of problems closed during the reporting period", "Problems moved to resolved or closed status this week"],
        ["", "New Problems Raised", "Number of new problem records opened this period", "Problems raised from incident trends or proactive analysis"],
        ["", "Avg Resolution Days (<5)", "Average calendar days from problem creation to closure", "Measured from problem record creation date to closure date"],
        ["", "Recurring Issues (<3)", "Number of issues recurring more than once in the period", "Incidents with same root cause appearing repeatedly; trend indicator"],
        ["", "Problems Aged >30 days (<5)", "Number of open problems older than 30 calendar days", "Aged problems requiring escalation and management review"],
        # Asset and Configuration
        ["Asset and Configuration", "Total CIs", "Total number of Configuration Items in ServiceNow CMDB", "All active CI records across all classes and environments"],
        ["", "CIs Verified This Week", "Number of CIs verified and validated during the reporting period", "CIs reviewed by CI owners and confirmed as accurate"],
        ["", "CI Accuracy (>95%)", "Percentage of CIs with accurate and up-to-date attribute data", "Verified through quarterly CMDB audit and weekly spot checks"],
        ["", "Orphan CIs (<10)", "Number of CIs with no parent relationship or service linkage", "CIs not linked to any business service or application in CMDB"],
        ["", "Stale Records (<20)", "Number of CI records not updated or validated in over 30 days", "Records requiring review and revalidation by CI owners"],
        ["", "Audit Compliance (>98%)", "Percentage compliance score from monthly CMDB audit", "Based on audit checklist covering completeness and accuracy"],
        # NOC
        ["NOC", "Alerts Received", "Total number of monitoring alerts received by NOC team", "Includes all severity levels from all monitored infrastructure"],
        ["", "Alerts Auto-Resolved", "Number of alerts automatically resolved without NOC intervention", "Auto-remediation scripts and rules resolving alerts without manual action"],
        ["", "P1 Incidents (<3)", "Number of Priority 1 incidents detected and managed by NOC", "Critical outages requiring immediate NOC escalation to MIM team"],
        ["", "P2 Incidents (<8)", "Number of Priority 2 incidents detected and managed by NOC", "High severity alerts requiring prompt NOC investigation and response"],
        ["", "MTTR Hours (<2)", "Mean Time To Restore service — average hours from alert to resolution", "Measured from alert trigger time to confirmed service restoration"],
        ["", "SLA Compliance (>95%)", "Percentage of incidents resolved within agreed SLA timeframe", "P1 < 4 hrs and P2 < 8 hrs resolution targets"],
        # ServiceNow
        ["ServiceNow", "Total Tickets", "Total service requests and incidents logged in ServiceNow", "Includes all request types and categories processed through the platform"],
        ["", "Tickets Resolved", "Number of tickets moved to resolved status in the period", "Resolved tickets pending user confirmation closure after 5 days"],
        ["", "Escalated Tickets", "Number of tickets escalated to L2 or L3 support teams", "Tickets that could not be resolved by first-line ServiceNow team"],
        ["", "Resolution Rate (>95%)", "Percentage of tickets resolved within agreed SLA target", "Calculated against resolution time targets by ticket category"],
        ["", "CSAT Score (>4.0 out of 5)", "Average customer satisfaction score from post-resolution surveys", "Automated survey sent to users upon ticket closure; scored 1-5"],
        ["", "First Call Resolution (>80%)", "Percentage of tickets resolved without requiring any reassignment", "Single-touch resolution without escalation to other teams or groups"],
        # Service Desk
        ["Service Desk", "Calls Received", "Total inbound contacts handled by Service Desk in the period", "Includes voice calls, chat sessions and email contacts"],
        ["", "Calls Resolved", "Total contacts resolved by first-line Service Desk agents", "Direct resolution without L2/L3 escalation required"],
        ["", "Abandoned Calls", "Number of contacts abandoned before being answered by an agent", "Key measure of queue management and staffing adequacy"],
        ["", "FCR Rate (>80%)", "Percentage of contacts resolved on first contact without escalation", "Primary measure of Service Desk effectiveness and knowledge quality"],
        ["", "Avg Handle Time Mins (<10)", "Average time in minutes to handle each contact end-to-end", "From contact start to documented resolution and system update"],
        ["", "SLA Compliance (>95%)", "Percentage of contacts responded to within agreed response times", "Based on agreed response targets by contact type and priority"],
    ]
    for row in cover_rows:
        ws_cover.append(row)

    # --- MIM (4 weekly rows — mirrors real scorecard format) ---
    ws_mim = wb.create_sheet("MIM")
    ws_mim.append([
        "Sprint No", "Total MI", "MIR Conducted", "# of P1", "# of P2",
        "100% Adherence to the MIM process (>99%)",
        "Accurate stakeholder communications (>99%)",
        "Consistency in the Service now fields (>99%)",
        "MIR Timeline (>99%)",
        "SLA compliance P1/P2 (>95%)",
        "SMS within 20 min P1/P2 (>99%)",
        "Email within 20 min engaged (>99%)",
        "Update cadence P1 30m P2 60m (>99%)",
        "MI whiteboard accuracy (>99%)",
        "Effective handover for RCA (>99%)",
        "Weekly Comments/Achievements",
    ])
    mim_data = [
        ("Sprint 11", 3, 3, 1, 2, 1.00, 1.00, 1.00, 1.00, 0.97, 1.00, 1.00, 1.00, 1.00, 1.00, None),
        ("Sprint 12", 4, 4, 1, 3, 1.00, 1.00, 1.00, 1.00, 0.96, 1.00, 1.00, 1.00, 1.00, 1.00, None),
        ("Sprint 13", 3, 3, 0, 3, 1.00, 1.00, 1.00, 1.00, 0.98, 1.00, 1.00, 1.00, 1.00, 1.00, None),
        ("Sprint 14", 2, 5, 0, 2, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00,
         "Achievements\nZero P1s this sprint\nMIR timeline 100% met across all incidents\nACL communications improved following ELT feedback\nConcerns\nSLA compliance dipped to 97% in Sprint 13, monitoring closely\nNext Week Focus\nContinue P1 prevention measures and runbook reviews\nReview ACL communication templates for ELT standards"),
    ]
    for i, row in enumerate(mim_data, start=2):
        sprint, tmi, mir, p1, p2 = row[0], row[1], row[2], row[3], row[4]
        kpis = row[5:15]
        comment = row[15]
        ws_mim.append([sprint, tmi, mir, p1, p2] + [None]*10 + [comment])
        for j, kv in enumerate(kpis, start=6):
            set_percent(ws_mim, ws_mim.cell(i, j), kv)

    # --- Change (4 weekly rows) ---
    ws_change = wb.create_sheet("Change")
    ws_change.append([
        "Week No", "Total Changes", "Emergency Changes",
        "Change Success Ratio (>99%)",
        "Change Causing MI (<5%)",
        "Failed Change Ratio (<1%)",
        "Unauthorised Change rate (<1%)",
        "Weekly Comments/Achievements",
    ])
    change_data = [
        ("Week 1", 48, 2, 1.00, 0.021, 0.000, 0.004, None),
        ("Week 2", 52, 1, 0.990, 0.019, 0.005, 0.003, None),
        ("Week 3", 45, 3, 1.00, 0.000, 0.000, 0.002, None),
        ("Week 4", 50, 2, 1.00, 0.000, 0.000, 0.006,
         "Achievements\nAll changes implemented within approved window\nZero P1 impact from changes this week\n2 emergency changes approved within 2 hrs of request\nConcerns\nUnauthorised change rate at 0.6%, approaching 1% threshold\nNext Week Focus\nReview change approval workflows for emergency category\nEnhance CAB pre-screening for high-risk changes"),
    ]
    for i, (wk, tot, emg, csr, cm, fcr, ucr, comment) in enumerate(change_data, start=2):
        ws_change.append([wk, tot, emg, None, None, None, None, comment])
        set_percent(ws_change, ws_change.cell(i, 4), csr)
        set_percent(ws_change, ws_change.cell(i, 5), cm)
        set_percent(ws_change, ws_change.cell(i, 6), fcr)
        set_percent(ws_change, ws_change.cell(i, 7), ucr)

    # --- Problem (4 weekly rows) ---
    ws_problem = wb.create_sheet("Problem")
    ws_problem.append([
        "Week No", "Open Problems", "Problems Resolved", "New Problems Raised",
        "Avg Resolution Days (<5)",
        "Recurring Issues (<3)",
        "Problems Aged >30 days (<5)",
        "Weekly Comments/Achievements",
    ])
    for i, (wk, op, pr, np, ard, ri, aged, comment) in enumerate([
        ("Week 1", 18, 8, 5, 3.8, 1, 3, None),
        ("Week 2", 15, 9, 4, 4.0, 2, 4, None),
        ("Week 3", 12, 8, 3, 3.5, 1, 3, None),
        ("Week 4", 11, 8, 3, 4.2, 2, 2,
         "Achievements\nClosed 3 aged problems this week\nRCA submitted for 2 P1 incidents on time\nProactive monitoring added for top 5 recurring incidents\nConcerns\nAvg resolution days increased to 4.2, approaching 5-day target\n2 recurring issues need additional root cause analysis\nNext Week Focus\nEscalate 2 remaining aged problems to senior management\nConduct root cause analysis workshop for recurring issues"),
    ], start=2):
        ws_problem.append([wk, op, pr, np, ard, ri, aged, comment])

    # --- Asset and Configuration (4 weekly rows + multigroup narrative columns) ---
    ws_cmdb = wb.create_sheet("Asset and Configuration")
    ws_cmdb.append([
        "Week No", "Total CIs", "CIs Verified This Week",
        "CI Accuracy (>95%)",
        "Orphan CIs (<10)",
        "Stale Records (<20)",
        "Audit Compliance (>98%)",
        "Group1_Name", "Group1_Achievements", "Group1_Concerns", "Group1_Focus",
        "Group2_Name", "Group2_Achievements", "Group2_Concerns", "Group2_Focus",
        "Group3_Name", "Group3_Achievements", "Group3_Concerns", "Group3_Focus",
        "Timeline",
        "Weekly Comments/Achievements",
    ])
    cmdb_kpi = [
        ("Week 1", 12480, 145, 0.978, 8, 14, 0.993),
        ("Week 2", 12495, 132, 0.980, 7, 13, 0.991),
        ("Week 3", 12510, 158, 0.975, 9, 15, 0.989),
        ("Week 4", 12524, 141, 0.977, 7, 12, 0.992),
    ]
    cmdb_narrative = [
        (None, None, None, None, None, None, None, None, None, None, None, None, None, None),
        (None, None, None, None, None, None, None, None, None, None, None, None, None, None),
        (None, None, None, None, None, None, None, None, None, None, None, None, None, None),
        (
            "CMDB",
            "Quarterly CMDB audit passed with 97.7% accuracy; 7 orphan CIs resolved this week; Azure VM CI class successfully onboarded",
            "15 stale records still pending owner review; CI reconciliation delayed for 2 business services",
            "Complete Azure VM reconciliation; Onboard Network team CI class to CMDB",
            "HAM",
            "Hardware asset register fully updated for the quarter; 4 end-of-life servers decommissioned and removed from register",
            "6 active assets without valid vendor support contracts; 3 assets missing location tags",
            "Complete EoL review for remaining server estate; Update procurement tracker with 2026 renewals",
            "SAM",
            "Software license compliance at 97%; 2 unused Oracle license sets reclaimed saving £18k annually",
            "Adobe Creative Cloud audit pending since last quarter; AWS reserved instance reconciliation overdue",
            "Submit Adobe audit results to procurement by Friday; Complete AWS reserved instance review",
            "Week 1: CMDB audit kickoff — baseline accuracy 97.5% | Week 2: HAM EoL review initiated — 12 servers flagged | Week 3: SAM license reclaim completed — 2 Oracle sets returned | Week 4: Quarterly audit report submitted to CIO; CMDB accuracy 97.7%",
            "CMDB audit passed. 7 orphan CIs scheduled for cleanup. Stale records reduced by 3 this week.",
        ),
    ]
    for i, ((wk, tot, ver, cia, oci, sr, ac), narr) in enumerate(
            zip(cmdb_kpi, cmdb_narrative), start=2):
        row_vals = [wk, tot, ver, None, oci, sr, None] + list(narr)
        ws_cmdb.append(row_vals)
        set_percent(ws_cmdb, ws_cmdb.cell(i, 4), cia)
        set_percent(ws_cmdb, ws_cmdb.cell(i, 7), ac)

    # --- NOC (4 weekly rows) ---
    ws_noc = wb.create_sheet("NOC")
    ws_noc.append([
        "Week No", "Alerts Received", "Alerts Auto-Resolved",
        "P1 Incidents (<3)",
        "P2 Incidents (<8)",
        "MTTR Hours (<2)",
        "SLA Compliance (>95%)",
        "Weekly Comments/Achievements",
    ])
    noc_data = [
        ("Week 1", 1842, 1560, 1, 5, 1.5, 0.971, None),
        ("Week 2", 1975, 1680, 2, 6, 1.8, 0.968, None),
        ("Week 3", 1723, 1490, 1, 7, 1.6, 0.972, None),
        ("Week 4", 1890, 1610, 2, 5, 1.7, 0.974,
         "Achievements\nAuto-resolution rate reached 85%, highest this quarter\nP1 incidents resolved in avg 1.4 hrs, well within 2-hr target\nNew runbook deployed for DB alert cluster reducing false positives\nConcerns\n2 P1 incidents detected this week, trend needs monitoring\nMTTR at 1.7 hrs for P2 incidents, slightly above 2-hr target\nNext Week Focus\nExpand auto-remediation scripts to cover network alert category\nReview P2 escalation path to reduce MTTR"),
    ]
    for i, (wk, ar, aar, p1, p2, mttr, sla, comment) in enumerate(noc_data, start=2):
        ws_noc.append([wk, ar, aar, p1, p2, mttr, None, comment])
        set_percent(ws_noc, ws_noc.cell(i, 7), sla)

    # --- ServiceNow (4 weekly rows) ---
    ws_snow = wb.create_sheet("ServiceNow")
    ws_snow.append([
        "Week No", "Total Tickets", "Tickets Resolved", "Escalated Tickets",
        "Resolution Rate (>95%)",
        "CSAT Score (>4.0 out of 5)",
        "First Call Resolution (>80%)",
        "Weekly Comments/Achievements",
    ])
    snow_data = [
        ("Week 1", 210, 201, 12, 0.957, 4.2, 0.81, None),
        ("Week 2", 225, 215, 10, 0.956, 4.3, 0.82, None),
        ("Week 3", 198, 192, 8, 0.970, 4.1, 0.83, None),
        ("Week 4", 230, 219, 9, 0.952, 4.4, 0.84,
         "Achievements\nCSAT improved to 4.4 out of 5, best score this quarter\nFCR consistently above 84% for third consecutive week\nEscalations reduced 25% versus prior week through knowledge base improvements\nConcerns\nResolution rate at 95.2%, borderline against 95% target\nNext Week Focus\nInvestigate root cause of borderline resolution rate\nPublish 10 additional knowledge articles for common request types"),
    ]
    for i, (wk, tot, res, esc, rr, csat, fcr, comment) in enumerate(snow_data, start=2):
        ws_snow.append([wk, tot, res, esc, None, csat, None, comment])
        set_percent(ws_snow, ws_snow.cell(i, 5), rr)
        set_percent(ws_snow, ws_snow.cell(i, 7), fcr)

    # --- Service Desk (4 weekly rows) ---
    ws_sd = wb.create_sheet("Service Desk")
    ws_sd.append([
        "Week No", "Calls Received", "Calls Resolved", "Abandoned Calls",
        "FCR Rate (>80%)",
        "Avg Handle Time Mins (<10)",
        "SLA Compliance (>95%)",
        "Weekly Comments/Achievements",
    ])
    sd_data = [
        ("Week 1", 320, 295, 18, 0.822, 8.5, 0.962, None),
        ("Week 2", 305, 280, 22, 0.818, 8.8, 0.958, None),
        ("Week 3", 340, 315, 15, 0.826, 8.2, 0.971, None),
        ("Week 4", 310, 288, 14, 0.829, 8.4, 0.965,
         "Achievements\nHandle time reduced to 8.4 mins against 10-min target\n12 new knowledge articles published this week\nAbandoned calls at 14, lowest point this month\nConcerns\nFCR at 82.9%, slightly below internal 85% target\nAbandoned calls still above sub-10 goal\nNext Week Focus\nDeploy new chatbot for tier-1 queries to improve FCR\nReview call routing rules to reduce abandonment rate"),
    ]
    for i, (wk, recv, resv, abn, fcr, aht, sla, comment) in enumerate(sd_data, start=2):
        ws_sd.append([wk, recv, resv, abn, None, aht, None, comment])
        set_percent(ws_sd, ws_sd.cell(i, 5), fcr)
        set_percent(ws_sd, ws_sd.cell(i, 7), sla)

    # --- Report_Config sheet ---
    ws_cfg = wb.create_sheet("Report_Config")
    from openpyxl.styles import Font, PatternFill
    hdr_fill = PatternFill(fill_type="solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF")
    # Max 2 charts per slide; columns 10-11 are the RAG threshold definitions
    headers_cfg = [
        "Team Name", "Chart 1 Type", "Chart 1 Columns",
        "Chart 2 Type", "Chart 2 Columns",
        "Slide Layout", "Summary Mode", "Include Insights",
        "Skip This Team", "Priority",
        "Green Threshold (%)", "Amber Threshold (%)",
    ]
    ws_cfg.append(headers_cfg)
    for cell in ws_cfg[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font

    # fmt: [team, c1type, c1cols, c2type, c2cols, layout, sum_mode, insights, skip, prio, green, amber]
    default_rows = [
        ["MIM",          "grouped_bar",    "auto", "line",          "auto", "standard",   "ai_write", "yes", "no", "normal", 99, 95],
        ["Change",       "bar_line_combo", "auto", "grouped_bar",   "auto", "standard",   "ai_write", "yes", "no", "normal", 99, 95],
        ["Problem",      "grouped_bar",    "auto", "line",          "auto", "standard",   "ai_write", "yes", "no", "normal", 95, 90],
        ["Asset and Configuration", "auto", "auto", "", "", "multigroup", "ai_write", "yes", "no", "normal", 95, 90],
        ["NOC",          "bar_dotted_line","auto", "grouped_bar",   "auto", "standard",   "ai_write", "yes", "no", "normal", 95, 90],
        ["Monitoring",   "auto",           "auto", "",              "",     "standard",   "ai_write", "yes", "no", "normal", 95, 90],
        ["ITOR",         "auto",           "auto", "",              "",     "standard",   "ai_write", "yes", "no", "normal", 95, 90],
        ["APAC SD",      "auto",           "auto", "",              "",     "standard",   "ai_write", "yes", "no", "normal", 80, 70],
        ["EMEA SD",      "auto",           "auto", "",              "",     "standard",   "ai_write", "yes", "no", "normal", 80, 70],
        ["ServiceNow",   "bar_line_combo", "auto", "line",          "auto", "standard",   "ai_write", "yes", "no", "normal", 95, 90],
        ["Service Desk", "grouped_bar",    "auto", "bar_dotted_line","auto","standard",   "ai_write", "yes", "no", "normal", 95, 90],
    ]
    for row in default_rows:
        ws_cfg.append(row)

    wb.save(excel_path)


def main():
    parser = argparse.ArgumentParser(description="ITSM Report Automation Agent — POC (Local Machine)")
    parser.add_argument("--run-once", action="store_true", help="Run the pipeline once and exit")
    parser.add_argument("--force", action="store_true",
                        help="Delete cached sheet hashes and today's PPTX before running, "
                             "forcing all slides to regenerate (use after code changes)")
    args = parser.parse_args()

    # Load env (optional).
    load_dotenv()

    repo_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(repo_dir, "config.yaml")
    config = _load_config(config_path)

    excel_watch_path = config.get("excel", {}).get("watch_path") or "./sample_data/sample_itsm.xlsx"
    excel_watch_path_abs = os.path.abspath(os.path.join(repo_dir, excel_watch_path)) if not os.path.isabs(excel_watch_path) else excel_watch_path

    if args.force:
        output_cfg = config.get("output", {})
        output_folder = output_cfg.get("folder", "./output")
        prefix = output_cfg.get("filename_prefix", "ITSM_Report")
        from datetime import datetime as _dt
        state_file = os.path.join(repo_dir, output_folder, f"{prefix}_sheet_state.json")
        today_pptx = os.path.join(repo_dir, output_folder,
                                  f"{prefix}_{_dt.now().strftime('%Y%m%d')}.pptx")
        for f in (state_file, today_pptx):
            if os.path.exists(f):
                try:
                    os.remove(f)
                    print(f"[--force] Deleted {f}")
                except PermissionError:
                    print(f"[--force] WARNING: '{f}' is open (locked). "
                          f"Close it in PowerPoint, then re-run. "
                          f"State was cleared — all sheets will regenerate.")

    _ensure_sample_excel_exists(excel_watch_path_abs)

    orchestrator = ReportOrchestrator(config)

    if args.run_once or args.force:
        result = orchestrator.run_pipeline(excel_watch_path_abs)
        print("\nPipeline Result:")
        for k in ["status", "slides_generated", "slides_skipped", "output_path", "duration_seconds"]:
            print(f"{k}: {result.get(k)}")
        return

    print(f"Watching Excel file for changes: {excel_watch_path_abs}")

    # Lazy import to keep startup fast.
    from tools.t1_file_watcher import ExcelFileWatcher

    def on_modified(_excel_path: str):
        result = orchestrator.run_pipeline(_excel_path)
        # Keep console output simple for POC.
        print(f"Pipeline finished: {result.get('status')} -> {result.get('output_path')}")

    watcher = ExcelFileWatcher(excel_watch_path_abs, on_modified)
    watcher.start()


if __name__ == "__main__":
    main()

