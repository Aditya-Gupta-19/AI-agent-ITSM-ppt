import os
import tempfile

import openpyxl
import pandas as pd
import pytest

from tools.t2_excel_reader import ExcelReader


def _write_workbook(wb, file_path: str) -> None:
    wb.save(file_path)


def _build_basic_workbook(sheet_names=("A", "B")) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws0 = wb.active
    ws0.title = sheet_names[0]
    ws0.append(["Col1", "Ratio (%)"])
    ws0.append([1, 0.5])

    for sn in sheet_names[1:]:
        ws = wb.create_sheet(sn)
        ws.append(["Col1", "Ratio (%)"])
        ws.append([2, 0.7])

    return wb


def test_reads_all_sheets():
    wb = _build_basic_workbook(("SheetA", "SheetB", "SheetC"))
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "test.xlsx")
        _write_workbook(wb, path)
        reader = ExcelReader()
        sheets = reader.read_all_sheets(path)
        assert set(sheets.keys()) == {"SheetA", "SheetB", "SheetC"}


def test_detects_empty_sheet():
    wb = _build_basic_workbook(("HasData",))
    wb.create_sheet("Empty")
    ws_empty = wb["Empty"]
    # Leave it truly empty (no rows).
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "test.xlsx")
        _write_workbook(wb, path)
        reader = ExcelReader()
        sheets = reader.read_all_sheets(path)
        assert sheets["Empty"]["is_empty"] is True


def test_column_type_detection():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "T"
    ws.append(["Ratio (%)", "Other"])
    ws.append([0.5, "x"])
    ws.append([0.7, "y"])

    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "test.xlsx")
        _write_workbook(wb, path)
        reader = ExcelReader()
        sheets = reader.read_all_sheets(path)
        column_types = sheets["T"]["column_types"]
        assert column_types["Ratio (%)"] == "percent_decimal"


def test_column_type_percent_decimal():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "T"
    ws.append(["SLA Compliance"])
    ws.append([0.95])
    ws.append([0.98])
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "test.xlsx")
        _write_workbook(wb, path)
        reader = ExcelReader()
        sheets = reader.read_all_sheets(path)
        assert sheets["T"]["column_types"]["SLA Compliance"] == "percent_decimal"


def test_get_user_summary_found():
    df = pd.DataFrame({"Weekly Comments/\nAchievements": [None, "Sprint 16: All KPIs met"]})
    reader = ExcelReader()
    assert reader.get_user_summary(df) == "Sprint 16: All KPIs met"


def test_get_user_summary_missing_column():
    df = pd.DataFrame({"KPI": [1, 2]})
    reader = ExcelReader()
    assert reader.get_user_summary(df) == ""


def test_get_user_summary_all_empty():
    df = pd.DataFrame({"Weekly Comments/Achievements": [None, None]})
    reader = ExcelReader()
    assert reader.get_user_summary(df) == ""


def test_detect_weekly_columns_stops_at_blank():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "T"
    ws.append(["Week", "KPI1", "KPI2", None, "Quarter", "Agg1"])
    reader = ExcelReader()
    indices = reader._detect_weekly_columns(ws)
    assert indices == [1, 2, 3]


def test_read_all_sheets_excludes_right_half():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MIM"
    ws.append(["Week", "KPI1", "KPI2", None, "Quarter"])
    ws.append([1, 0.5, 0.8, None, 3])
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "test.xlsx")
        _write_workbook(wb, path)
        reader = ExcelReader()
        sheets = reader.read_all_sheets(path)
        df = sheets["MIM"]["dataframe"]
        assert df.shape[1] == 3


def test_read_report_config_returns_dict():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report_Config"
    ws.append(["Team Name","Chart 1 Type","Chart 1 Columns","Chart 2 Type",
                "Chart 2 Columns","Slide Layout","Summary Mode",
                "Include Insights","Skip This Team","Priority"])
    ws.append(["MIM","grouped_bar","auto","none","auto","standard",
               "ai_write","yes","no","normal"])
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "test.xlsx")
        _write_workbook(wb, path)
        reader = ExcelReader()
        config = reader.read_report_config(path)
        assert "MIM" in config
        mim = config["MIM"]
        assert mim["charts"] == [
            {"type": "grouped_bar", "columns": "auto"},
            {"type": "none", "columns": "auto"},
        ]
        assert mim["layout"] == "standard"
        assert mim["summary_mode"] == "ai_write"
        assert mim["include_insights"] is True
        assert mim["skip"] is False
        assert mim["priority"] == "normal"


def test_read_report_config_no_sheet_returns_empty():
    wb = _build_basic_workbook(("SheetA",))
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "test.xlsx")
        _write_workbook(wb, path)
        reader = ExcelReader()
        assert reader.read_report_config(path) == {}


def test_handles_merged_cells():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Merged"
    ws.append(["A", "B"])
    ws.append([None, None])

    ws["A2"].value = "MergedVal"
    ws.merge_cells("A2:B2")

    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "test.xlsx")
        _write_workbook(wb, path)
        reader = ExcelReader()
        sheets = reader.read_all_sheets(path)
        df = sheets["Merged"]["dataframe"]
        assert df.iloc[0]["A"] == "MergedVal"
        assert df.iloc[0]["B"] == "MergedVal"

