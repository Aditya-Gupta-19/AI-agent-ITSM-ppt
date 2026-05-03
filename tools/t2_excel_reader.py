import io
import os
import tempfile
from dataclasses import dataclass
from typing import Any, Dict, List, Optional

import openpyxl
import pandas as pd


@dataclass
class SheetData:
    dataframe: pd.DataFrame
    headers: list[str]
    row_count: int
    is_empty: bool
    column_types: Dict[str, str]


class ExcelReader:
    def __init__(self, config: Optional[dict] = None):
        self.kpi_sheet_name = None
        if config:
            self.kpi_sheet_name = (
                config.get("excel", {}).get("kpi_sheet_name")
                or config.get("excel", {}).get("kpi_sheet")
            )

    def read_all_sheets(self, file_path: str) -> Dict[str, Dict[str, Any]]:
        """
        Returns:
        {
          "sheet_name": {
            "dataframe": pd.DataFrame,
            "headers": list[str],
            "row_count": int,
            "is_empty": bool,
            "column_types": dict  # {col_name: "numeric" | "percent" | "text"}
          }
        }
        """

        with open(file_path, "rb") as f:
            file_bytes = io.BytesIO(f.read())
        wb = openpyxl.load_workbook(file_bytes, data_only=True)

        # Unmerge merged cells and forward-fill from the top-left cell.
        for ws in wb.worksheets:
            merged_ranges = list(ws.merged_cells.ranges)
            for merged in merged_ranges:
                # merged can be cast to str like "A1:B2"
                min_row = merged.min_row
                min_col = merged.min_col
                value = ws.cell(row=min_row, column=min_col).value
                ws.unmerge_cells(str(merged))
                for r in range(merged.min_row, merged.max_row + 1):
                    for c in range(merged.min_col, merged.max_col + 1):
                        ws.cell(row=r, column=c).value = value

        with tempfile.TemporaryDirectory() as d:
            tmp_path = os.path.join(d, "merged_unpacked.xlsx")
            wb.save(tmp_path)
            wb.close()

            sheets: Dict[str, Dict[str, Any]] = {}
            with pd.ExcelFile(tmp_path, engine="openpyxl") as xl:
                for sheet_name in xl.sheet_names:
                    header_mode = None
                    if self.kpi_sheet_name and sheet_name == self.kpi_sheet_name:
                        # KPI sheet is parsed positionally (A/B/C); treat first row as data.
                        header_mode = None  # header=None
                    else:
                        header_mode = 0  # header row is row 1

                    # Detect weekly (left-half) columns for team sheets only.
                    usecols = None
                    is_special = (
                        (self.kpi_sheet_name and sheet_name == self.kpi_sheet_name)
                        or sheet_name == "Report_Config"
                    )
                    if not is_special and sheet_name in wb.sheetnames:
                        weekly_indices = self._detect_weekly_columns(wb[sheet_name])
                        if len(weekly_indices) < len(list(wb[sheet_name][1])):
                            usecols = [i - 1 for i in weekly_indices]

                    if header_mode is None:
                        df = pd.read_excel(tmp_path, sheet_name=sheet_name, header=None,
                                           engine="openpyxl", usecols=usecols)
                    else:
                        df = pd.read_excel(tmp_path, sheet_name=sheet_name, header=0,
                                           engine="openpyxl", usecols=usecols)

                    # Drop fully empty rows.
                    df = df.dropna(how="all")

                    # Normalize headers as strings for downstream consistency.
                    headers = [str(h) for h in df.columns.tolist()] if len(df.columns) else []

                    # Detect empty sheet.
                    is_empty = df.empty or df.shape[0] == 0
                    row_count = int(df.shape[0])

                    column_types: Dict[str, str] = {}
                    if not is_empty:
                        for col in df.columns:
                            col_name = str(col)
                            column_types[col_name] = self._detect_column_type(df[col])

                    sheets[sheet_name] = {
                        "dataframe": df,
                        "headers": headers,
                        "row_count": row_count,
                        "is_empty": bool(is_empty),
                        "column_types": column_types,
                    }

            return sheets

    @staticmethod
    def parse_weekly_comments(text: str) -> dict:
        """
        Parses the Weekly Comments/Achievements cell text into 3 named sections.
        Recognises headers: Achievements, Concerns, Next Week Focus (case-insensitive).
        Returns {"achievements": [str], "concerns": [str], "focus": [str]}.
        Content before any header lands in 'achievements' by default.
        """
        result: dict = {"achievements": [], "concerns": [], "focus": []}
        if not text or not str(text).strip():
            return result

        lines = str(text).strip().replace("\r\n", "\n").replace("\r", "\n").split("\n")

        # Ordered: longest keywords first so "next week focus" beats "next week"
        _HEADER_MAP = [
            (["next week focus", "next week", "next steps", "action items",
              "upcoming", "actions for next week"], "focus"),
            (["achievement", "achievements", "this week achievements",
              "what went well", "wins", "accomplishment"], "achievements"),
            (["concern", "concerns", "risk", "risks", "issue", "issues",
              "challenge", "challenges", "corrective action", "preventive action",
              "areas for improvement", "comments for amber"], "concerns"),
        ]

        current = "achievements"

        for line in lines:
            stripped = line.strip()
            if not stripped:
                continue
            clean = stripped.lstrip("•-→✓▸►*#\t0123456789.) ")
            lower = clean.lower().rstrip(".:- ").strip()

            if len(lower) <= 55:
                found = False
                for keywords, section in _HEADER_MAP:
                    if any(lower == kw or lower.startswith(kw) for kw in keywords):
                        current = section
                        found = True
                        break
                if found:
                    continue

            content = stripped.lstrip("•-→✓▸►*#\t") .strip()
            if content:
                result[current].append(content)

        return result

    @staticmethod
    def parse_multigroup_comments(text: str) -> dict:
        """
        Parses multi-group narrative text (e.g. CMDB/HAM/SAM in one cell).
        Groups are detected by short non-bullet lines that are NOT standard section headers.
        Each group may contain Achievements / Concerns / Next Week Focus sub-sections.
        Returns {"groups": [{"name":str, "achievements":[str], "concerns":[str], "focus":[str]}],
                 "timeline": str}.
        """
        if not text or not str(text).strip():
            return {"groups": [], "timeline": ""}

        lines = str(text).strip().replace("\r\n", "\n").replace("\r", "\n").split("\n")

        _SECTION_HEADERS = {
            "achievement": "achievements", "achievements": "achievements",
            "this week": "achievements",
            "concern": "concerns", "concerns": "concerns",
            "next week focus": "focus", "next week": "focus", "focus": "focus",
            "next steps": "focus", "action items": "focus",
        }
        _TIMELINE_KW = {"timeline", "key updates", "timeline & key updates",
                        "timeline and key updates"}

        groups: list = []
        timeline_lines: list = []
        current_group: dict = {}
        current_section: str = "achievements"
        in_timeline = False

        def _new_group(name: str) -> dict:
            g: dict = {"name": name, "achievements": [], "concerns": [], "focus": []}
            groups.append(g)
            return g

        for line in lines:
            stripped = line.strip()
            if not stripped:
                continue
            clean = stripped.lstrip("•-→✓▸►*#\t0123456789.) ")
            lower = clean.lower().rstrip(".:- ").strip()

            # Timeline section
            if any(lower == kw or lower.startswith(kw) for kw in _TIMELINE_KW):
                in_timeline = True
                continue
            if in_timeline:
                content = stripped.lstrip("•-→✓▸►*#\t").strip()
                if content:
                    timeline_lines.append(content)
                continue

            # Section header within a group
            if len(lower) <= 55:
                matched_sec = None
                for kw, sec in _SECTION_HEADERS.items():
                    if lower.startswith(kw):
                        matched_sec = sec
                        break
                if matched_sec:
                    current_section = matched_sec
                    continue

                # Potential new group header: short, no leading bullet/number
                if not stripped[0] in "•-→✓▸►*":
                    # Check it doesn't look like plain sentence content
                    if not any(lower.endswith(end) for end in [".", ":", ";"]) or len(lower) < 20:
                        if not current_group or lower != current_group.get("name", "").lower():
                            current_group = _new_group(clean)
                            current_section = "achievements"
                            continue

            # Content line
            content = stripped.lstrip("•-→✓▸►*#\t").strip()
            if not content:
                continue
            if not current_group:
                current_group = _new_group("Group 1")
            current_group[current_section].append(content)

        return {
            "groups": groups,
            "timeline": " | ".join(timeline_lines) if timeline_lines else "",
        }

    def read_multigroup_data(self, df: pd.DataFrame) -> dict:
        """
        Extracts 3-subgroup narrative data from a DataFrame.
        Looks for columns named Group1_Name, Group1_Achievements, Group1_Concerns,
        Group1_Focus (and Group2_*, Group3_*) plus a Timeline column.
        Returns {"groups": [...], "timeline": "..."}.
        """
        groups = []
        for gi in range(1, 4):
            def _find_col(suffix, _gi=gi):
                target = f"Group{_gi}_{suffix}"
                return next((c for c in df.columns if str(c).strip() == target), None)

            name_col = _find_col("Name")
            ach_col  = _find_col("Achievements")
            con_col  = _find_col("Concerns")
            foc_col  = _find_col("Focus")

            if name_col is None and ach_col is None:
                continue

            def _last(col):
                if col is None:
                    return ""
                s = df[col].dropna()
                return str(s.iloc[-1]).strip() if not s.empty else ""

            groups.append({
                "name":         _last(name_col) or f"Group {gi}",
                "achievements": _last(ach_col),
                "concerns":     _last(con_col),
                "focus":        _last(foc_col),
            })

        tl_col = next((c for c in df.columns if str(c).strip() == "Timeline"), None)
        timeline = ""
        if tl_col is not None:
            s = df[tl_col].dropna()
            if not s.empty:
                timeline = str(s.iloc[-1]).strip()

        return {"groups": groups, "timeline": timeline}

    def get_user_summary(self, df: pd.DataFrame) -> str:
        """
        Finds the 'Weekly Comments/Achievements' column and returns
        the last non-empty value as a string. Returns "" if not found.
        """
        comment_col = None
        for col in df.columns:
            col_lower = str(col).lower()
            if "comment" in col_lower and "achievement" in col_lower:
                comment_col = col
                break

        if comment_col is None:
            return ""

        series = df[comment_col].dropna()
        if series.empty:
            return ""

        return str(series.iloc[-1]).strip()

    def _detect_weekly_columns(self, ws) -> list:
        """
        Returns 1-based column indices belonging to the LEFT (weekly) half only.
        Strategy: once 3+ non-blank header columns are seen, stop at the first
        blank column. If no blank gap is found, return all columns.
        """
        headers = []
        for cell in ws[1]:  # row 1
            headers.append(cell.value)

        result = []
        data_seen = 0
        for idx, h in enumerate(headers, start=1):
            is_blank = h is None or str(h).strip() == ""
            if is_blank:
                if data_seen >= 3:
                    break  # right-half starts here
                # skip leading blanks
            else:
                data_seen += 1
                result.append(idx)

        return result if result else [1]

    def read_report_config(self, file_path: str) -> dict:
        """
        Reads the 'Report_Config' sheet from the Excel workbook.
        Returns {} if the sheet is not found.
        Column order (1-indexed): Team Name | Chart1 Type | Chart1 Cols |
          Chart2 Type | Chart2 Cols | Chart3 Type | Layout | Summary Mode |
          Include Insights | Skip | Priority
        """
        try:
            with open(file_path, "rb") as fh:
                buf = io.BytesIO(fh.read())
            wb = openpyxl.load_workbook(buf, data_only=True)
        except Exception:
            return {}

        try:
            if "Report_Config" not in wb.sheetnames:
                return {}

            ws = wb["Report_Config"]
            result: dict = {}

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue

                team_name = str(row[0]).strip()
                if not team_name:
                    continue

                def _cell(idx: int, _row=row) -> str:
                    val = _row[idx] if idx < len(_row) else None
                    return str(val).strip().lower() if val is not None else ""

                def _bool(idx: int, _row=row) -> bool:
                    val = _row[idx] if idx < len(_row) else None
                    return str(val).strip().lower() == "yes" if val is not None else False

                charts: List[dict] = []
                for type_idx, col_idx in [(1, 2), (3, 4)]:
                    ctype = _cell(type_idx)
                    ccols = _cell(col_idx)
                    if ctype:
                        charts.append({"type": ctype, "columns": ccols or "auto"})
                # Chart 3 type only (no separate columns column)
                ctype3 = _cell(5)
                if ctype3:
                    charts.append({"type": ctype3, "columns": "auto"})

                def _float_val(idx: int, default: float = 0.0) -> float:
                    val = row[idx] if idx < len(row) else None
                    try:
                        return float(val) if val is not None else default
                    except (TypeError, ValueError):
                        return default

                include_raw = _cell(8)
                result[team_name] = {
                    "charts": charts,
                    "layout": _cell(6) or "standard",
                    "summary_mode": _cell(7) or "ai_write",
                    "include_insights": (include_raw != "no") if include_raw else True,
                    "skip": _bool(9),
                    "priority": _cell(10) or "normal",
                    "green_threshold": _float_val(11, 95.0),
                    "amber_threshold": _float_val(12, 90.0),
                }

            return result
        finally:
            wb.close()

    def read_kpi_definitions(self, file_path: str) -> dict:
        """
        Reads KPI definitions from Cover Page (4-column format):
          Col A: Team | Col B: KPI Name | Col C: Definition | Col D: Sub-Definition
        Returns {team_name: [{"kpi_name": ..., "definition": ..., "sub_definition": ...}]}
        Returns {} if Cover Page not found or has fewer than 4 columns.
        """
        try:
            with open(file_path, "rb") as fh:
                buf = io.BytesIO(fh.read())
            wb = openpyxl.load_workbook(buf, data_only=True)
        except Exception:
            return {}

        cover_name = None
        for name in wb.sheetnames:
            if "cover" in name.lower():
                cover_name = name
                break
        if not cover_name:
            wb.close()
            return {}

        ws = wb[cover_name]
        result: dict = {}
        current_team = ""

        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row:
                continue
            col_a = str(row[0]).strip() if row[0] is not None else ""
            col_b = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            col_c = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            col_d = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""

            if col_a:
                current_team = col_a
            if not current_team or not col_b:
                continue
            # Skip header row
            if col_b.lower() in ("kpi name", "kpi / metric", "metric"):
                continue
            if current_team not in result:
                result[current_team] = []
            result[current_team].append({
                "kpi_name": col_b,
                "definition": col_c,
                "sub_definition": col_d,
            })

        wb.close()
        return result

    def _detect_column_type(self, series: pd.Series) -> str:
        non_null = series.dropna()
        if non_null.empty:
            return "text"

        numeric = pd.to_numeric(non_null, errors="coerce").dropna()
        if numeric.empty:
            return "text"

        try:
            min_v = float(numeric.min())
            max_v = float(numeric.max())
        except Exception:
            return "text"

        if 0.0 <= min_v <= 1.0 and 0.0 <= max_v <= 1.0:
            # Keyword check: column represents a percentage stored as 0-1 decimal.
            _PCT_KEYWORDS = [
                "ratio", "rate", "accuracy", "compliance", "adherence",
                "percent", "fcr", "sla", "timeline", "whiteboard",
                "handover", "stakeholder", "consistency", "resolution",
                "cadence", "within", "sms", "email", "effective",
            ]
            col_lower = str(series.name).lower() if hasattr(series, "name") else ""
            is_pct_keyword = "%" in col_lower or any(kw in col_lower for kw in _PCT_KEYWORDS)
            if is_pct_keyword:
                return "percent_decimal"
            return "percent"

        return "numeric"

